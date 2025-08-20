# invoices/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, F
from django.utils.dateparse import parse_date
from django.utils import timezone
from products.models import Product
from customers.models import Customers
from .models import Invoice, InvoiceItem, DailySale
from datetime import datetime, date, timedelta
from django.http import HttpResponse # JsonResponse - not used directly in this specific view snippet
from django.conf import settings
from django.contrib import messages
from django.template.loader import render_to_string, get_template
from django.core.mail import EmailMessage


from io import BytesIO
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def create_invoice_view(request):
    products_for_js = Product.objects.filter(status=True)

    if request.method == 'POST':
        customer_phone = request.POST.get('customer_phone', '').strip()
        customer_name = request.POST.get('customer_name', '').strip()

        if not customer_phone: customer_phone = '+910000000000' # Default or validate as required
        if not customer_name: customer_name = 'Walk-in Customer'

        payment_method = request.POST.get('payment_method', 'Cash')
        given_amount_str = request.POST.get('given_amount', '0')

        try:
            given_amount_float = float(given_amount_str) if given_amount_str else 0.0
        except ValueError:
            messages.error(request, "Invalid given amount. Please enter a valid number.")
            return render(request, 'create_invoice.html', {
                'products': products_for_js, 'customer_phone': customer_phone,
                'customer_name': customer_name, 'payment_method': payment_method,
                'given_amount_str': given_amount_str
            })

        customer, created = Customers.objects.get_or_create(
            Phone_number=customer_phone,
            defaults={'name': customer_name}
        )
        if not created and customer.name != customer_name and customer_name.lower() != 'walk-in customer':
            customer.name = customer_name
            customer.save()

        product_ids = request.POST.getlist('product_id[]')
        quantities_str = request.POST.getlist('quantity[]')
        item_prices_str = request.POST.getlist('item_price[]') # Unit price for the specific item/size/weight
        item_size_names = request.POST.getlist('item_size_name[]')
        # item_custom_weights_str = request.POST.getlist('item_custom_weight[]') # This will only contain values for custom cakes

        if not product_ids:
            messages.error(request, "No products were selected. Please add items to the bill.")
            return render(request, 'create_invoice.html', {'products': products_for_js})

        if not (len(product_ids) == len(quantities_str) == len(item_prices_str) == len(item_size_names)):
            messages.error(request, "There was a mismatch in submitted product data lengths. Please try again.")
            return render(request, 'create_invoice.html', {
                'products': products_for_js, 'customer_phone': customer_phone,
                'customer_name': customer_name, 'payment_method': payment_method,
                'given_amount': given_amount_str
            })

        calculated_total_invoice_float = 0.0
        invoice_items_to_create = []

        # Keep track of which custom weights have been used if item_custom_weight[] is sparse
        custom_weight_iter = iter(request.POST.getlist('item_custom_weight[]'))

        for i in range(len(product_ids)):
            pid = product_ids[i]
            qty_s = quantities_str[i]
            price_s = item_prices_str[i] 
            size_name = item_size_names[i]
            
            custom_weight_for_item = None # For optional storage in InvoiceItem
            
            try:
                product_obj = Product.objects.get(id=pid)
                quantity = int(qty_s)
                actual_item_unit_price_float = float(price_s)

                # Check if this is a custom weight cake based on product category and size_name format
                if "cake" in product_obj.category.name.lower() and "kg" in size_name.lower() and any(char.isdigit() for char in size_name):
                    try:
                        # If the frontend sends item_custom_weight[] for every item (even if empty for non-cakes), this direct indexing would work.
                        # However, the current JS only sends it for custom cakes. So, we use an iterator.
                        custom_weight_for_item = float(next(custom_weight_iter))
                    except (StopIteration, ValueError, TypeError):
                        # Fallback: try to parse from size_name if iterator fails or value is bad
                        # This might happen if form submission is manipulated or JS changes.
                        try:
                            weight_part = size_name.lower().replace("kg", "").strip()
                            custom_weight_for_item = float(weight_part)
                        except ValueError:
                            custom_weight_for_item = None # Still couldn't determine
                        if custom_weight_for_item is None:
                             print(f"Warning: Could not determine custom weight for cake item: {product_obj.name}, size: {size_name}")


                if quantity <= 0:
                    messages.error(request, f"Quantity for {product_obj.name} ({size_name}) must be positive.")
                    return render(request, 'create_invoice.html', {'products': products_for_js})

                line_item_total_float = actual_item_unit_price_float * quantity
                calculated_total_invoice_float += line_item_total_float
                
                item_data_to_append = {
                    'product': product_obj,
                    'quantity': quantity,
                    'price_float': actual_item_unit_price_float,
                    'total_float': line_item_total_float,
                    'size_name': size_name,
                    # 'custom_weight_val': custom_weight_for_item # Store if you have a field in InvoiceItem
                }
                invoice_items_to_create.append(item_data_to_append)

            except Product.DoesNotExist:
                messages.error(request, f"Product with ID {pid} not found. Please try again.")
                return render(request, 'create_invoice.html', {'products': products_for_js})
            except (ValueError, TypeError) as e:
                messages.error(request, f'Error processing item: Product ID {pid}, Price "{price_s}", Size "{size_name}", Qty "{qty_s}". Details: {e}')
                return render(request, 'create_invoice.html', {'products': products_for_js})
        
        final_total_for_invoice_int = round(calculated_total_invoice_float)
        given_amount_int = round(given_amount_float)

        if payment_method == 'Cash' and given_amount_int < final_total_for_invoice_int:
            messages.error(request, f"Given amount (₹{given_amount_int}) is less than total amount (₹{final_total_for_invoice_int}).")
            return render(request, 'create_invoice.html', {
                'products': products_for_js, 'customer_phone': customer_phone,
                'customer_name': customer_name, 'payment_method': payment_method,
                'given_amount': given_amount_str,
                'calculated_total_display': final_total_for_invoice_int,
                'error_message_custom': 'Given amount too low.'
            })

        with transaction.atomic():
            invoice = Invoice.objects.create(
                customer=customer,
                total_amount=final_total_for_invoice_int,
                grand_total=final_total_for_invoice_int,
                payment_method=payment_method
            )
            request.session['given_amount_for_receipt'] = given_amount_int if payment_method == 'Cash' else final_total_for_invoice_int
            request.session['balance_return_for_receipt'] = (given_amount_int - final_total_for_invoice_int) if payment_method == 'Cash' else 0

            for item_data in invoice_items_to_create:
                InvoiceItem.objects.create(
                    invoice=invoice,
                    product=item_data['product'],
                    quantity=item_data['quantity'],
                    price=round(item_data['price_float']), # This is the price of one unit of this item
                    total=round(item_data['total_float']),
                    size_name=item_data['size_name']
                    # If you have a custom_weight field in InvoiceItem model:
                    # custom_weight=item_data.get('custom_weight_val') # or however you stored it
                )
            
            sale_date_obj = date.today()
            total_items_sold_count = sum(item['quantity'] for item in invoice_items_to_create)

            daily_sale, created = DailySale.objects.get_or_create(
                date=sale_date_obj,
                defaults={ # These defaults are used only if 'created' is True
                    'total_amount': 0,
                    'total_orders': 0,
                    'total_items_sold': 0
                }
            )
            daily_sale.total_amount = F('total_amount') + final_total_for_invoice_int
            daily_sale.total_orders = F('total_orders') + 1
            daily_sale.total_items_sold = F('total_items_sold') + total_items_sold_count
            daily_sale.save()


        messages.success(request, f"Invoice #{invoice.id} created successfully!")
        return redirect('invoice_detail', pk=invoice.id)

    return render(request, 'create_invoice.html', {'products': products_for_js})


def invoice_detail_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    items = invoice.items.all()

    given_amount = request.session.pop('given_amount_for_receipt', invoice.grand_total)
    balance_return = request.session.pop('balance_return_for_receipt', 0)

    return render(request, 'invoice_detail.html', {
        'invoice': invoice,
        'items': items,
        'given_amount': given_amount,
        'balance_return': balance_return
    })


def invoice_list(request):
    invoices = Invoice.objects.all().order_by('-date')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    phone = request.GET.get('phone')
    amount_str = request.GET.get('amount')

    if from_date_str:
        from_date_obj = parse_date(from_date_str)
        if from_date_obj: invoices = invoices.filter(date__date__gte=from_date_obj)
    if to_date_str:
        to_date_obj = parse_date(to_date_str)
        if to_date_obj: invoices = invoices.filter(date__date__lte=to_date_obj)
    if phone:
        invoices = invoices.filter(customer__Phone_number__icontains=phone)
    if amount_str:
        try:
            amount_float = float(amount_str)
            invoices = invoices.filter(grand_total__gte=round(amount_float))
        except ValueError:
            messages.warning(request, "Invalid amount format for filtering.")

    context = {
        'invoices': invoices, 'from_date': from_date_str, 'to_date': to_date_str,
        'phone': phone, 'amount': amount_str,
    }
    return render(request, 'invoice_list.html', context)


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    print(f"PDF Generation Error: {pdf.err} for template {template_src}")
    return None


def daily_sales_report(request):
    selected_date_str = request.GET.get('date')
    filter_by = request.GET.get('filter', 'daily')
    export_pdf = request.GET.get('export_pdf')
    export_xls = request.GET.get('export_xls')

    invoices_qs = Invoice.objects.none()
    total_amount_val = 0
    product_sales_agg = []
    report_period_display = "No date selected"
    start_dt, end_dt = None, None
    current_error_message = None

    today_date_obj = timezone.now().date() # Use timezone.now().date() for timezone-aware 'today'
    today_start_dt = datetime.combine(today_date_obj, datetime.min.time())
    today_end_dt = datetime.combine(today_date_obj, datetime.max.time())

    todays_invoices_qs = Invoice.objects.filter(date__range=(today_start_dt, today_end_dt))
    todays_aggregation = todays_invoices_qs.aggregate(total=Sum('grand_total'))
    todays_total_amount_val = todays_aggregation['total'] if todays_aggregation['total'] is not None else 0

    todays_product_sales_agg = (
        InvoiceItem.objects.filter(invoice__in=todays_invoices_qs)
        .values('product__name', 'size_name') # Keep size_name if relevant
        .annotate(total_quantity=Sum('quantity'), total_value=Sum('total'))
        .order_by('product__name', 'size_name', '-total_quantity')
    )

    if selected_date_str:
        try:
            date_obj = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format. Use YYYY-MM-DD.')
            current_error_message = 'Invalid date format. Use YYYY-MM-DD.'
            # Prepare minimal context for error display
            context = {
                'selected_date': selected_date_str,
                'filter_by': filter_by,
                'error_message': current_error_message,
                # Include today's sales data even on error for the selected date
                'todays_date_display': today_date_obj.strftime('%d %b %Y'),
                'todays_total_sales': todays_total_amount_val,
                'todays_product_sales': todays_product_sales_agg,
                'has_todays_sales': todays_total_amount_val > 0,
                'invoices': invoices_qs, # empty
                'product_sales': product_sales_agg, # empty
                'total_amount': total_amount_val # 0
            }
            return render(request, 'daily_sales.html', context)

        if filter_by == 'daily':
            start_dt = datetime.combine(date_obj, datetime.min.time())
            end_dt = datetime.combine(date_obj, datetime.max.time())
            report_period_display = f"{date_obj.strftime('%d %b %Y')}"
        
        elif filter_by == 'weekly':
            start_of_week = date_obj - timedelta(days=date_obj.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            start_dt = datetime.combine(start_of_week, datetime.min.time())
            end_dt = datetime.combine(end_of_week, datetime.max.time())
            report_period_display = f"Week: {start_of_week.strftime('%d %b %Y')} - {end_of_week.strftime('%d %b %Y')}"
        
        elif filter_by == 'monthly':
            start_of_month = date_obj.replace(day=1)
            next_month_start = (start_of_month + timedelta(days=32)).replace(day=1)
            end_of_month = next_month_start - timedelta(days=1)
            start_dt = datetime.combine(start_of_month, datetime.min.time())
            end_dt = datetime.combine(end_of_month, datetime.max.time())
            report_period_display = f"Month: {start_of_month.strftime('%B %Y')}"
        
        elif filter_by == 'yearly':
            start_of_year = date_obj.replace(month=1, day=1)
            end_of_year = date_obj.replace(month=12, day=31)
            start_dt = datetime.combine(start_of_year, datetime.min.time())
            end_dt = datetime.combine(end_of_year, datetime.max.time())
            report_period_display = f"Year: {start_of_year.strftime('%Y')}"
        
        if start_dt and end_dt:
            invoices_qs = Invoice.objects.filter(date__range=(start_dt, end_dt)).order_by('date')
            aggregation = invoices_qs.aggregate(total=Sum('grand_total'))
            total_amount_val = aggregation['total'] if aggregation['total'] is not None else 0
            product_sales_agg = (
                InvoiceItem.objects.filter(invoice__in=invoices_qs)
                .values('product__name', 'size_name')
                .annotate(total_quantity=Sum('quantity'), total_value=Sum('total'))
                .order_by('product__name', 'size_name', '-total_quantity')
            )

    context = {
        'invoices': invoices_qs, 'selected_date': selected_date_str,
        'total_amount': total_amount_val, 'filter_by': filter_by,
        'product_sales': product_sales_agg, 'report_period_display': report_period_display,
        'start_date_filter': start_dt.strftime('%Y-%m-%d') if start_dt else '',
        'end_date_filter': end_dt.strftime('%Y-%m-%d') if end_dt else '',
        'error_message': current_error_message,
        'todays_date_display': today_date_obj.strftime('%d %b %Y'),
        'todays_total_sales': todays_total_amount_val,
        'todays_product_sales': todays_product_sales_agg,
        'has_todays_sales': todays_total_amount_val > 0
    }

    if export_pdf == 'true' and selected_date_str and start_dt and end_dt:
        pdf_template_path = 'reports/daily_sales_pdf.html' # Assumed path
        pdf = render_to_pdf(pdf_template_path, context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"sales_report_{filter_by}_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        else:
            messages.error(request, "Error generating PDF report.")
            query_params = request.GET.copy()
            if 'export_pdf' in query_params: del query_params['export_pdf']
            return redirect(request.path_info + '?' + query_params.urlencode())

    if export_xls == 'true' and selected_date_str and start_dt and end_dt:
        wb = openpyxl.Workbook()
        ws_invoices = wb.active; ws_invoices.title = "Invoices"
        ws_invoices['A1'] = f"Sales Report - {filter_by.capitalize()}"; ws_invoices['A1'].font = Font(bold=True, size=16)
        ws_invoices['A2'] = f"Period: {report_period_display}"
        ws_invoices['A3'] = f"Total Sales Amount: ₹{total_amount_val:,.2f}"; ws_invoices['A3'].font = Font(bold=True)
        
        invoice_headers = ["S.No", "Bill No", "Customer Name", "Customer Phone", "Date & Time", "Payment Method", "Grand Total (₹)"]
        for col, title in enumerate(invoice_headers, 1):
            cell = ws_invoices.cell(row=5, column=col, value=title)
            cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')

        for i, inv in enumerate(invoices_qs, 1):
            ws_invoices.cell(row=i+5, column=1, value=i)
            ws_invoices.cell(row=i+5, column=2, value=inv.id)
            ws_invoices.cell(row=i+5, column=3, value=inv.customer.name if inv.customer else "N/A")
            ws_invoices.cell(row=i+5, column=4, value=inv.customer.Phone_number if inv.customer else "N/A")
            ws_invoices.cell(row=i+5, column=5, value=inv.date.strftime("%d %b %Y %H:%M"))
            ws_invoices.cell(row=i+5, column=6, value=inv.payment_method)
            ws_invoices.cell(row=i+5, column=7, value=inv.grand_total).number_format = '#,##0'
        
        for col_idx in range(1, ws_invoices.max_column + 1):
            max_len = max(len(str(cell.value)) for cell in ws_invoices[get_column_letter(col_idx)] if cell.value)
            ws_invoices.column_dimensions[get_column_letter(col_idx)].width = (max_len + 2) * 1.1

        if product_sales_agg:
            ws_products = wb.create_sheet(title="Product Sales")
            ws_products['A1'] = "Product Sales Summary"; ws_products['A1'].font = Font(bold=True, size=14)
            prod_headers = ["Product Name", "Size/Weight", "Total Qty Sold", "Total Value (₹)"]
            for col, title in enumerate(prod_headers, 1):
                cell = ws_products.cell(row=3, column=col, value=title)
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
            for i, item in enumerate(product_sales_agg, 1):
                ws_products.cell(row=i+3, column=1, value=item['product__name'])
                ws_products.cell(row=i+3, column=2, value=item.get('size_name', 'Standard'))
                ws_products.cell(row=i+3, column=3, value=item['total_quantity'])
                ws_products.cell(row=i+3, column=4, value=item['total_value']).number_format = '#,##0'
            for col_idx in range(1, ws_products.max_column + 1):
                max_len = max(len(str(cell.value)) for cell in ws_products[get_column_letter(col_idx)] if cell.value)
                ws_products.column_dimensions[get_column_letter(col_idx)].width = (max_len + 2) * 1.1
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        response = HttpResponse(excel_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename_xls = f"sales_report_{filter_by}_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename_xls}"'
        return response

    return render(request, 'daily_sales.html', context)


def send_simple_daily_report_email(request):
    if request.method == 'POST':
        report_date_obj = date.today()
        start_dt = datetime.combine(report_date_obj, datetime.min.time())
        end_dt = datetime.combine(report_date_obj, datetime.max.time())

        total_sales_data = Invoice.objects.filter(date__range=(start_dt, end_dt)).aggregate(
            total_amount_sum=Sum('grand_total')
        )
        overall_total_amount_today = total_sales_data.get('total_amount_sum') or 0
        has_sales = overall_total_amount_today > 0

        email_subject = f"Daily Sales Summary - {getattr(settings, 'SHOP_NAME', 'Our Shop')} - {report_date_obj.strftime('%d/%m/%Y')}"
        email_context = {
            'report_date': report_date_obj,
            'overall_total_amount': overall_total_amount_today,
            'shop_name': getattr(settings, 'SHOP_NAME', 'Our Shop'),
            'has_sales': has_sales,
        }
        html_message = render_to_string('reports/daily_sales_report_email.html', email_context) # Assumed path
        
        owner_email = getattr(settings, 'OWNER_EMAIL_ADDRESS', None)
        if not owner_email:
            messages.error(request, "Owner email address not configured in settings.")
            return redirect('daily_sales_report')

        try:
            email = EmailMessage(email_subject, html_message, settings.DEFAULT_FROM_EMAIL, [owner_email])
            email.content_subtype = "html"
            email.send()
            messages.success(request, f"Today's sales summary sent to {owner_email}!")
        except Exception as e:
            print(f"EMAIL SENDING ERROR: {e}")
            messages.error(request, f"Failed to send email summary: {e}")
        
        return redirect('daily_sales_report')
    return redirect('daily_sales_report')